#!/usr/bin/env python3
"""Build 14-backlog.xlsx — Jira/ADO-import-ready backlog workbook (source: 14-backlog.md)."""
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
NAVY, TEAL, LINE = "12303F", "0FA68E", "D9E2DE"
PRI_FILL = {"P0":"FBEAE8","P1":"FDF3E3","P2":"E7F5F1"}
PRI_INK  = {"P0":"B3362B","P1":"A8690A","P2":"0B7D6B"}

EPICS = {
 "E1":"Production Safety & Operability","E2":"Compliance: DPDP, Form C, GST",
 "E3":"OTA Ingestion completeness","E4":"Money correctness","E5":"Notifications & Messaging activation",
 "E6":"Auth & access hardening","E7":"Fleet & onboarding","E8":"Localization & accessibility",
 "E9":"Verification & spec-truth",
}
# epic, id, role, story, AC, refs, pri, pts
S = [
("E1","US-101","MindBit ops","Automated daily backups with an offsite copy per client DB","GIVEN a client deployment WHEN the nightly job runs THEN a restorable snapshot exists offsite, retention 30d, and failure alerts ops.","GAP-1, RSK-08","P0",5),
("E1","US-102","MindBit ops","Documented, drilled restore runbook","GIVEN a snapshot WHEN the runbook is executed on a scratch instance quarterly THEN RTO ≤ 4h is verified and logged.","GAP-1","P0",3),
("E1","US-103","MindBit ops","Error tracking + uptime checks on every deployment","GIVEN an unhandled error or downtime WHEN it occurs THEN it appears in monitoring within 5 min tagged by client, and the UI shows an incident ID.","GAP-17","P0",5),
("E1","US-104","Owner","Sync health visible (iCal last-success, forwarder heartbeat, cron status) with staleness warnings","GIVEN a feed not synced >12h WHEN I open Feeds/Today THEN a visible warning shows the age AND a push/alert fires (with E5).","GAP-5, BR-AVL-05","P0",5),
("E1","US-105","MindBit ops","Fleet dashboard: version, health, cron, backup status across clients","GIVEN ≥2 deployments WHEN I open the dashboard THEN each shows version+health+last backup+last sync, red on breach.","GAP-18","P1",8),
("E2","US-201","Owner","Pre-filled Form C artefact for a foreign guest, with submission tracking","GIVEN a checked-in foreign guest WHEN I tap 'Form C' THEN a print-ready artefact with the 13 fields renders; a due-in-24h reminder persists until 'submitted' is ticked (audited).","GAP-7, Q-LEG-02","P0",5),
("E2","US-202","Owner","Export and erase all data held about a guest (DPDP rights)","GIVEN a guest requests erasure WHEN I trigger it THEN PII is removed/anonymised, bookings keep non-identifying financial integrity, and the action is audited.","GAP-8","P0",8),
("E2","US-203","MindBit","Sane ID-scan retention default; purge provably runs","GIVEN a new property WHEN created THEN idRetentionDays defaults (e.g. 180); purge logs deletions; owner notified of policy.","BR-GST-05","P0",3),
("E2","US-204","MindBit","Breach-notification runbook + guest-facing privacy notice (incl. AI/US-cloud processing)","Docs approved by counsel; notice linked in app.","NFR-PRV-01/04","P0",3),
("E2","US-205","Owner (GST)","Invoices carry sequential per-FY numbers, GST lines, stored immutably","GIVEN GSTIN configured WHEN an invoice is issued THEN an Invoice record exists with number series, tax lines per rate, immutable snapshot; reprint is identical.","GAP-11, Q-FIN-02","P1",8),
("E2","US-206","Guest","Server-side PDF invoices","Pixel-stable PDF, downloadable on a phone.","GAP-11","P2",5),
("E3","US-301","Reception","OTA modification/cancellation emails match the existing booking and guide the update","GIVEN a booking with otaRef X WHEN a modification email for X is ingested THEN the pending item links to the booking, shows a diff, and 'Apply' updates via the normal conflict-checked path.","GAP-2","P0",8),
("E3","US-302","MindBit","Fixture corpus of real OTA emails gates parser changes in CI","GIVEN the corpus (Q-OTA-02) WHEN the parser changes THEN CI runs fixtures and reports extraction accuracy.","RSK-13","P1",5),
("E3","US-303","Owner","Configurable iCal sync frequency (up to hourly)","Configurable; respects OTA politeness; health surfaced (US-104).","GAP-6","P2",3),
("E3","US-304","Owner","Last-room oversell buffer per room type","GIVEN buffer=1 on Deluxe(3 units) WHEN 2 are booked THEN agent/community availability reports 0; direct booking still possible with warning.","GAP-24","P2",5),
("E3","US-305","Reception","Duplicate OTA confirmations never create two bookings","GIVEN a booking with otaRef X WHEN the same-ref email is ingested THEN it is flagged duplicate; no new pending create.","BR-OTA-03","P1",3),
("E3","US-306","MindBit","Card data redacted from ingested emails before storage","GIVEN a fixture containing a virtual card WHEN ingested THEN the stored record contains none of it.","ASM-18, Q-OTA-03","P1",2),
("E4","US-401","Engineering","Money as integer paise end-to-end with a typed money util","All money fields migrated; property tests on arithmetic; CSVs/invoices unchanged to the rupee.","GAP-9","P0",8),
("E4","US-402","Owner","Staff API responses never contain money fields","GIVEN a reception session WHEN any endpoint returns a reservation THEN grossAmount etc. are absent; contract tests per role.","GAP-12","P0",5),
("E4","US-403","Engineering","Postgres RLS enforces property scoping (defence-in-depth)","RLS policies on tenant tables keyed to session property; raw-SQL leak test proves containment.","GAP-12, RSK-17","P1",8),
("E4","US-404","Owner","Payment corrections are void+reversal with audit","Edit disabled; void creates a reversing entry; both audited.","BR-PAY-05, Q-FIN-04","P1",3),
("E4","US-405","Owner","OTA payout recording with owed-vs-received per OTA","Payout entity matched to bookings; Finance shows variance.","GAP-13","P2",5),
("E4","US-406","Owner","Deterministic refund rules per ratified answers","Ladder base per Q-FIN-07; refund ≤ paid enforced; OTA-collect path documented.","BR-CANC-03/04/05","P1",3),
("E4","US-407","Owner","Razorpay capture + room-hold (when keys provided)","Idempotent webhook (replay-safe test); hold expires → release; exclusion constraint untouched.","FR-PAY-5","P1",8),
("E5","US-501","Owner","Push notification for high-severity escalations and new conflicts","GIVEN a subscribed device WHEN the event fires THEN push within 1 min, deep-linking to the item; per-event toggles in Settings.","GAP-14, BR-AI-05","P0",5),
("E5","US-502","MindBit","WhatsApp templates (En/Kha/Hi) approved and parameterised","Templates submitted+approved; outbox sends real messages with params; language chosen per guest.","GAP-3","P1",8),
("E5","US-503","Owner","Guest replies visible in a unified message log","Inbound webhook stores replies; thread view per guest; ownership per Q-AI-04 decision.","GAP-4","P2",5),
("E5","US-504","Guest","Message opt-out and quiet hours","STOP handling; no sends 22:00–07:00 except OTP-class.","RSK-15","P1",3),
("E6","US-601","Staff","Join via email invite and set own password","Invite → expiring link → set password → role/property pre-assigned.","GAP-10","P0",5),
("E6","US-602","Any user","Self-service password reset","Reset link, single-use, rate-limited.","GAP-10","P0",3),
("E6","US-603","Owner","Disable/role change revokes sessions","Per Q-SEC-03 ratified semantics; test proves stale session rejected.","AMB-21","P1",3),
("E6","US-604","Owner","ID-document access is logged","Every view/download → audit entry; owner-visible.","Q-SEC-04, GAP-15","P1",2),
("E6","US-605","Engineering","Keyed hashing for shared-list phone matching","Pepper per network; enumeration infeasible; migration for existing hashes.","NFR-SEC-09","P1",3),
("E6","US-606","Engineering","CI guard for raw-SQL tenant scoping","Lint rule / grep-gate fails PRs with unscoped tenant-table raw SQL.","NFR-SEC-06","P1",2),
("E7","US-701","MindBit ops","One-command client provisioning (deploy+DB+seed+env)","New deployment live <2h hands-on; checklist auto-verified.","GAP-18, Q-DEP-01","P1",8),
("E7","US-702","New owner","Setup wizard: property, rooms, channels, policies, staff","First-run wizard completes to a bookable state without support.","GAP-18","P1",8),
("E7","US-703","MindBit ops","Staged fleet upgrades (canary → all) with pre-migration backup gates","Scripted; halt-on-failure; version visible in fleet dashboard.","RSK-09","P1",5),
("E7","US-704","Client","Complete data export on request (offboarding)","DB export + documents + CSVs delivered; runbook.","GAP-23","P2",3),
("E7","US-705","MindBit","Pilot billing operates with a rate card","Pricing ratified (Q-BUS-03); invoices issued; revenue tracked vs grant milestone.","GAP-27","P1",2),
("E7","US-706","Sales","Seeded demo tenant","Demo data realistic; reset script.","GAP-30","P2",2),
("E8","US-801","Engineering","All UI strings externalized (i18n framework)","No hardcoded user-facing strings; language-switch scaffolding.","GAP-16","P1",8),
("E8","US-802","Khasi-speaking operator","Core flows available in Khasi","Today/Calendar/Booking/Payments/Housekeeping translated; reviewed per Q-L10N-03.","GAP-16","P2",8),
("E8","US-803","Any user","Core flows meet WCAG 2.1 AA basics","Contrast, touch targets ≥44px, labels; audit checklist passes.","GAP-21","P2",5),
("E9","US-901","Team","Offline behaviour audited and specified honestly","Doc: which reads cache, which writes queue, conflict UX; User Guide corrected if overstated; 2G field test.","GAP-25, Q-TEC-04","P0",5),
("E9","US-902","Team","Community topology documented as an ADR","Where registries live, who controls, failure modes; counsel-reviewed.","GAP-26, Q-LEG-03","P0",3),
("E9","US-903","Engineering","Route-level test: booking create overlap → 409","Vitest route test green; guards db-errors sniffing.","NFR-MNT-02","P0",2),
("E9","US-904","Engineering","Housekeeping derivation test","Checkout ⇒ task appears; mark-clean ⇒ ready.","NFR-MNT-02","P1",2),
("E9","US-905","Team","Room-assignment product stance ratified and recorded","Decision workshop output; if type-level chosen, spike estimates constraint redesign.","GAP-29, Q-OPS-05","P1",2),
("E9","US-906","Engineering","Per-feed iCal tokens replace the single token","Rotation per feed; old URLs invalidated gracefully.","ROADMAP suggestion","P2",3),
("E9","US-907","Engineering","Guest merge tool with audit","Merge preserves stays/LTV; irreversible warning; audited.","GAP-19","P2",5),
]

wb = Workbook()
thin = Border(*[Side(style="thin", color=LINE)]*4)
def header(ws, row, cells):
    for c,(txt,w) in enumerate(cells,1):
        cell = ws.cell(row=row,column=c,value=txt)
        cell.font = Font(name="Arial",bold=True,color="FFFFFF",size=10)
        cell.fill = PatternFill("solid",fgColor=NAVY)
        cell.alignment = Alignment(vertical="center",wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(c)].width = w

# ---- Sheet 1: Backlog ----
ws = wb.active; ws.title = "Backlog"
ws.sheet_view.showGridLines = False
ws["A1"] = "Guest House Ops Hub — Product Backlog v1.0"
ws["A1"].font = Font(name="Arial",bold=True,size=14,color=NAVY)
ws["A2"] = "Source: discovery doc 14 (14-backlog.md). Delta-to-enterprise-quality; delivered baseline excluded. Cross-refs: doc 02 (FR/NFR), 05 (BR), 06 (GAP), 07 (RSK), 09 (Q)."
ws["A2"].font = Font(name="Arial",size=9,color="5B6B64")
header(ws,4,[("Epic",6),("Epic name",26),("Story ID",9),("As a…",16),("Story",44),
             ("Acceptance criteria (Given/When/Then)",62),("References",20),("Priority",9),("Points",8),("Status",12)])
r = 5
for e,sid,role,story,ac,refs,pri,pts in S:
    vals = [e,EPICS[e],sid,role,story,ac,refs,pri,pts,"Proposed"]
    for c,v in enumerate(vals,1):
        cell = ws.cell(row=r,column=c,value=v)
        cell.font = Font(name="Arial",size=9.5)
        cell.alignment = Alignment(vertical="top",wrap_text=(c in (2,4,5,6,7)))
        cell.border = thin
    p = ws.cell(row=r,column=8)
    p.fill = PatternFill("solid",fgColor=PRI_FILL[pri]); p.font = Font(name="Arial",size=9.5,bold=True,color=PRI_INK[pri])
    p.alignment = Alignment(horizontal="center",vertical="top")
    ws.cell(row=r,column=9).alignment = Alignment(horizontal="center",vertical="top")
    r += 1
ws.auto_filter.ref = f"A4:J{r-1}"
ws.freeze_panes = "A5"

# ---- Sheet 2: Epic summary (formulas) ----
es = wb.create_sheet("Epic Summary"); es.sheet_view.showGridLines = False
es["A1"] = "Epic Summary — counts & points derive from the Backlog sheet"
es["A1"].font = Font(name="Arial",bold=True,size=13,color=NAVY)
header(es,3,[("Epic",7),("Epic name",34),("Stories",9),("Points",9),("P0 stories",10),("P1 stories",10),("P2 stories",10),("P0 points",10)])
rr = 4
last = r-1
for e,name in EPICS.items():
    es.cell(row=rr,column=1,value=e); es.cell(row=rr,column=2,value=name)
    es.cell(row=rr,column=3,value=f'=COUNTIF(Backlog!$A$5:$A${last},A{rr})')
    es.cell(row=rr,column=4,value=f'=SUMIF(Backlog!$A$5:$A${last},A{rr},Backlog!$I$5:$I${last})')
    es.cell(row=rr,column=5,value=f'=COUNTIFS(Backlog!$A$5:$A${last},A{rr},Backlog!$H$5:$H${last},"P0")')
    es.cell(row=rr,column=6,value=f'=COUNTIFS(Backlog!$A$5:$A${last},A{rr},Backlog!$H$5:$H${last},"P1")')
    es.cell(row=rr,column=7,value=f'=COUNTIFS(Backlog!$A$5:$A${last},A{rr},Backlog!$H$5:$H${last},"P2")')
    es.cell(row=rr,column=8,value=f'=SUMIFS(Backlog!$I$5:$I${last},Backlog!$A$5:$A${last},A{rr},Backlog!$H$5:$H${last},"P0")')
    for c in range(1,9):
        cell = es.cell(row=rr,column=c); cell.font = Font(name="Arial",size=10); cell.border = thin
        if c>=3: cell.alignment = Alignment(horizontal="center")
    rr += 1
es.cell(row=rr,column=2,value="TOTAL").font = Font(name="Arial",bold=True,size=10)
for c,col in ((3,"C"),(4,"D"),(5,"E"),(6,"F"),(7,"G"),(8,"H")):
    cell = es.cell(row=rr,column=c,value=f'=SUM({col}4:{col}{rr-1})')
    cell.font = Font(name="Arial",bold=True,size=10); cell.border = thin
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid",fgColor="E7F5F1")
es.cell(row=rr,column=1).border = thin; es.cell(row=rr,column=2).border = thin

# ---- Sheet 3: DoR / DoD / legend ----
dd = wb.create_sheet("DoR · DoD · Legend"); dd.sheet_view.showGridLines = False
dd.column_dimensions["A"].width = 26; dd.column_dimensions["B"].width = 110
rows = [
 ("Definition of Ready","Open questions referenced by the story are answered; acceptance criteria agreed; test data identified; any touch to the GiST-constraint/migration machinery flagged for senior review."),
 ("Definition of Done","Code + tests (unit; route-level where API changes) green in CI; migration via db:migrate:new only; docs updated (USER-GUIDE/ARCHITECTURE); audit/RBAC respected; deployed to a pilot-mirror environment; demoed."),
 ("Priority P0","Production-safety / legal — gate for scaling past first pilots (Roadmap Phase 1)."),
 ("Priority P1","Scale & commercial — grant Q2/Q3 alignment (Roadmap Phase 2)."),
 ("Priority P2","Quality & experience — Roadmap Phase 3."),
 ("Status values","Proposed → Ready → In progress → Done (set 'Proposed' by default; update during grooming)."),
 ("Import note","Backlog sheet columns map 1:1 to Jira CSV import (Epic Link, Summary=Story, Description=AC, Labels=References, Priority, Story Points)."),
]
dd["A1"] = "Working agreements"; dd["A1"].font = Font(name="Arial",bold=True,size=13,color=NAVY)
rr = 3
for k,v in rows:
    a = dd.cell(row=rr,column=1,value=k); b = dd.cell(row=rr,column=2,value=v)
    a.font = Font(name="Arial",bold=True,size=10); b.font = Font(name="Arial",size=10)
    a.border = thin; b.border = thin
    b.alignment = Alignment(wrap_text=True,vertical="top"); a.alignment = Alignment(vertical="top")
    rr += 1

wb.save(ROOT/"14-backlog.xlsx")
print("saved", len(S), "stories")
